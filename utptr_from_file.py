import utptr_classes
import openpyxl
import utptr_log as log

fileName = 'input/input.xlsx'

def readDevs(firstRow = 4, lastRow = 28):
    devsArray = []
    wb = openpyxl.load_workbook(fileName)
    if "Разработчики" in wb.get_sheet_names():
        ws = wb.get_sheet_by_name("Разработчики")
        colDevId = ws['A']
        colDevType = ws['B']
        colDevName = ws['C']
        colDevHoursPrimary = ws['D']
        colDevHoursSecondary = ws['E']
        colDevHoursExcess = ws['F']

        devsArray = []

        def notNumberToZero(val):
            if not isinstance(val, int) and not isinstance(val, float):
                return 0
            else:
                return val

        for i in range(firstRow, lastRow):
            devsArray.append(utptr_classes.Quota(colDevId[i].value,
                                                 colDevType[i].value,
                                                 colDevName[i].value,
                                                 notNumberToZero(colDevHoursPrimary[i].value),
                                                 notNumberToZero(colDevHoursSecondary[i].value),
                                                 notNumberToZero(colDevHoursExcess[i].value)))
            log.dev(colDevId[i].value, 'created with primary hours', notNumberToZero(colDevHoursPrimary[i].value))
            log.dev(colDevId[i].value, 'created with secondary hours', notNumberToZero(colDevHoursSecondary[i].value))
            log.dev(colDevId[i].value, 'created with extra hours', notNumberToZero(colDevHoursExcess[i].value))

    else:
        print("Нет листа с разработчиками")

    return devsArray